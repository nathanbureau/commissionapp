import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import re
import io
import random
import time as _time

REP_CONFIG = {
    "Joe Marshall":       {"level": "AE1", "region": "AUS", "currency": "AUD"},
    "Jack Eagle":         {"level": "AE1", "region": "UK",  "currency": "GBP"},
    "Jay Hudon":          {"level": "AE1", "region": "CAN", "currency": "CAD"},
    "Kevin Craig":        {"level": "AE1", "region": "CAN", "currency": "CAD"},
    "John Yienger":       {"level": "AE1", "region": "US TX", "currency": "USD"},
    "Tyler Givens":       {"level": "AE1", "region": "US NY", "currency": "USD"},
    "Brandon Brown":      {"level": "AE1", "region": "CAN", "currency": "CAD"},
    "Brett Robinson":     {"level": "AE1", "region": "US TX", "currency": "USD"},
    "Connor Harper":      {"level": "AE1", "region": "US NY", "currency": "USD"},
    "Jett Laws":          {"level": "AE1", "region": "US TX", "currency": "USD"},
    "Branson Wilson":     {"level": "AE1", "region": "US TX", "currency": "USD"},
    "Peter Holt":         {"level": "AE1", "region": "CAN", "currency": "CAD"},
    "Reuben Zuidhof":     {"level": "AE1", "region": "CAN", "currency": "CAD"},
    "Sam Quennell":       {"level": "AE1", "region": "AUS", "currency": "AUD"},
    "Adam Morgan":        {"level": "AE1", "region": "US NY", "currency": "USD"},
    "Grace Aicardi":      {"level": "AE1", "region": "AUS", "currency": "AUD"},
    "Anton Weininger":    {"level": "AE1", "region": "UK",  "currency": "GBP"},
    "Nik Balashov":       {"level": "AE1", "region": "UK",  "currency": "GBP"},
    "Harm Magis":         {"level": "AE1", "region": "UK",  "currency": "GBP"},
    "Megan Scholz":       {"level": "AE1", "region": "AUS", "currency": "AUD"},
    "Alex Kretowicz":     {"level": "AE1", "region": "CAN", "currency": "CAD"},
    "Harry Steele":        {"level": "AE2", "region": "AUS", "currency": "AUD"},
    "Joshua Cherry":       {"level": "AE2", "region": "CAN", "currency": "CAD"},
    "Lachie Topp":         {"level": "AE2", "region": "UK",  "currency": "GBP"},
    "Ryan Lenz":           {"level": "AE2", "region": "US NY", "currency": "USD"},
    "Kyle Harms":          {"level": "AE2", "region": "CAN", "currency": "CAD"},
    "Alex DeRenzis":           {"level": "CAM", "region": "CAN", "currency": "CAD"},
    "Graeme Hodson-Walker":    {"level": "CAM", "region": "CAN", "currency": "CAD"},
    "Natasha Lewis":           {"level": "CAM", "region": "US NY", "currency": "USD"},
    "Ella Horner":             {"level": "CAM", "region": "AUS", "currency": "AUD"},
    "Grace Randell":           {"level": "CAM", "region": "AUS", "currency": "AUD"},
    "Halle Smith":             {"level": "CAM", "region": "US TX", "currency": "USD"},
    "Ragan Sims":              {"level": "CAM", "region": "US TX", "currency": "USD"},
    "Kyle McCulloch":          {"level": "CAM", "region": "US NY", "currency": "USD"},
    "Danielle Celentano":      {"level": "CAM", "region": "CAN", "currency": "CAD"},
    "Rachelle Sampson":        {"level": "CAM", "region": "CAN", "currency": "CAD"},
    "Kathryn Nicholson-Brown": {"level": "CAM", "region": "UK",  "currency": "GBP"},
    "Nicole Murphy":           {"level": "CAM", "region": "CAN", "currency": "CAD"},
    "Marta Menendez":          {"level": "CAM", "region": "UK",  "currency": "GBP"},
    "Marcus De Verteuil":      {"level": "Manager", "region": "CAN", "currency": "CAD"},
    "Geddes Carrington":       {"level": "Manager", "region": "US",  "currency": "USD"},
}

QB_NAME_MAP = {"Josh Cherry": "Joshua Cherry", "Marcus de Verteuil": "Marcus De Verteuil"}
OWNER_NAME_MAP = {"Jett (Phillip) Laws": "Jett Laws"}
DEFAULT_RATES = {"USD": 1.0, "CAD": 0.74, "AUD": 0.65, "GBP": 1.26, "EUR": 1.08}
CURRENCY_SYMBOLS = {"USD": "$", "CAD": "C$", "AUD": "A$", "GBP": "\u00a3", "EUR": "\u20ac"}

def clean_money(val):
    try:
        if pd.isna(val): return 0.0
        s = str(val).replace(',', '').replace('$', '').replace('\u00a3', '').replace('A$', '').replace('C$', '').strip()
        if s.startswith('(') and s.endswith(')'): return -float(s[1:-1])
        return float(s)
    except Exception:
        return 0.0

def clean_inv_key(val):
    return "".join(re.findall(r'\d+', str(val)))

def clean_owner(name):
    if pd.isna(name): return ""
    name = re.sub(r'\s*\(Deactivated User\)', '', str(name).strip())
    if name in OWNER_NAME_MAP: return OWNER_NAME_MAP[name]
    return name.split('(')[0].strip()

def convert_currency(amount, src, dst, rates):
    if pd.isna(amount) or amount == 0: return 0.0
    if src == dst: return round(float(amount), 2)
    usd = float(amount) * rates.get(src, 1.0)
    target_rate = rates.get(dst, 1.0)
    return round(usd / target_rate, 2) if target_rate else 0.0

def fmt(val, currency="USD"):
    try: v = float(val)
    except (ValueError, TypeError): return str(val)
    sym = CURRENCY_SYMBOLS.get(currency, "$")
    return f"{sym}{v:,.2f}"

def get_commission_rate(rep, channel):
    cfg = REP_CONFIG.get(rep)
    if not cfg: return 0.0
    lv, rg = cfg['level'], cfg['region']
    ch = str(channel).lower().strip() if pd.notna(channel) else "inbound"
    outbound = "outbound" in ch
    if rep == "Marcus De Verteuil": return 0.05 if outbound else 0.02
    if lv == "Manager": return 0.0
    if lv == "CAM": return 0.01 if outbound else 0.0
    if any(x in ch for x in ("dealer", "architect", "rfp", "tender")): return 0.01
    ret = any(x in ch for x in ("return", "expansion", "customer"))
    if lv == "AE1":
        if rg in ("CAN", "UK", "US TX"): return 0.05 if outbound else 0.02
        if rg in ("AUS", "US NY"): return 0.05 if outbound else 0.015
    if lv == "AE2":
        if rg == "AUS": return 0.05 if outbound else (0.03 if ret else 0.015)
        if rg in ("CAN", "US TX"): return 0.05 if outbound else (0.03 if ret else 0.02)
        if rg == "US NY": return 0.06 if outbound else (0.025 if ret else 0.02)
        if rg == "UK": return 0.05 if outbound else (0.02 if ret else 0.015)
    return 0.0

def get_bonus_thresholds(rep):
    cfg = REP_CONFIG.get(rep)
    if not cfg: return [500000]
    lv, rg = cfg['level'], cfg['region']
    if rep == "Harry Steele": return [288000, 324000, 360000, 396000, 432000, 468000]
    if lv == "AE1":
        if rg == "CAN": return [273600, 307800, 342000, 376200, 410400, 444600]
        if rg == "UK": return [180000, 202500, 225000, 247500, 270000, 292500]
        if rg in ("US NY", "US TX"): return [302400, 340200, 378000, 415800, 453600, 491400]
        if rg == "AUS": return [259200, 291600, 324000, 356400, 388800, 421200]
    if lv == "AE2":
        if rg in ("AUS", "CAN"): return [364800, 410400, 456000, 501600, 547200, 592800]
        if rg in ("US NY", "US TX"): return [391200, 440100, 489000, 537900, 586800, 635700]
        if rg == "UK": return [300000, 337500, 375000, 412500, 450000, 487500]
    if lv == "CAM":
        if rg == "UK": return [360000, 405000, 450000, 495000, 540000, 585000]
        return [500000, 562500, 625000, 687500, 750000, 812500]
    return [500000, 562500, 625000, 687500, 750000, 812500]

def _get_tiers(rep):
    cfg = REP_CONFIG.get(rep)
    if not cfg: return []
    lv, rg = cfg['level'], cfg['region']
    if rep == "Harry Steele":
        return [(288000,0.002),(324000,0.005),(360000,0.012),(396000,0.014),(432000,0.018),(468000,0.020)]
    if lv == "AE1":
        if rg == "CAN": return [(273600,0.002),(307800,0.005),(342000,0.014),(376200,0.016),(410400,0.018),(444600,0.020)]
        if rg == "AUS": return [(259200,0.002),(291600,0.005),(324000,0.012),(356400,0.014),(388800,0.018),(421200,0.020)]
        if rg == "UK": return [(180000,0.002),(202500,0.005),(225000,0.014),(247500,0.016),(270000,0.018),(292500,0.020)]
        return [(302400,0.002),(340200,0.005),(378000,0.014),(415800,0.016),(453600,0.018),(491400,0.020)]
    if lv == "AE2":
        if rg == "AUS": return [(364800,0.002),(410400,0.005),(456000,0.012),(501600,0.014),(547200,0.018),(592800,0.020)]
        if rg == "CAN": return [(364800,0.002),(410400,0.005),(456000,0.014),(501600,0.018),(547200,0.020),(592800,0.022)]
        if rg == "UK": return [(300000,0.002),(337500,0.005),(375000,0.014),(412500,0.018),(450000,0.020),(487500,0.022)]
        return [(391200,0.002),(440100,0.005),(489000,0.014),(537900,0.018),(586800,0.020),(635700,0.022)]
    if lv == "CAM":
        if rg == "UK": return [(360000,0.002),(405000,0.004),(450000,0.017),(495000,0.019),(540000,0.021),(585000,0.022)]
        if rg == "CAN": return [(500000,0.010),(562500,0.0125),(625000,0.015),(687500,0.0175),(750000,0.020),(812500,0.0225)]
        if rg == "US NY": return [(500000,0.010),(562500,0.0125),(625000,0.015),(687500,0.0175),(750000,0.020),(812500,0.0225)]
        if rg == "US TX": return [(500000,0.015),(562500,0.0175),(625000,0.020),(687500,0.0225),(750000,0.025),(812500,0.0275)]
        if rg == "AUS": return [(500000,0.0075),(562500,0.010),(625000,0.0125),(687500,0.015),(750000,0.0175),(812500,0.020)]
    return []

def _get_tier_rate(attainment, rep):
    tiers = _get_tiers(rep)
    if not tiers or attainment < tiers[0][0]: return 0.0
    rate = 0.0
    for threshold, r in tiers:
        if attainment >= threshold: rate = r
        else: break
    return rate

def get_tier_name(attainment, rep):
    thresholds = get_bonus_thresholds(rep)
    labels = ["Below 80%", "80-90%", "90-100%", "100-110%", "110-120%", "120-130%", "130%+"]
    if attainment < thresholds[0]: return labels[0]
    for i, t in enumerate(thresholds):
        if attainment < t: return labels[i] if i < len(labels) else f"Tier {i}"
    return labels[-1]

def parse_qb_sales_detail(file_obj):
    try:
        fname = file_obj.name if hasattr(file_obj, 'name') else str(file_obj)
        raw = pd.read_csv(file_obj, header=None) if fname.endswith('.csv') else pd.read_excel(file_obj, header=None)
    except Exception:
        return pd.DataFrame()
    raw = raw.iloc[:, :min(10, len(raw.columns))].copy()
    ncols = len(raw.columns)
    header_idx = 4
    for i in range(min(20, len(raw))):
        row_vals = [str(raw.iloc[i, j]).strip().lower() if pd.notna(raw.iloc[i, j]) else '' for j in range(ncols)]
        if any(x in row_vals for x in ['date', 'transaction date']) and '#' in row_vals:
            header_idx = i
            break
    header_vals = [str(raw.iloc[header_idx, j]).strip().lower() if pd.notna(raw.iloc[header_idx, j]) else '' for j in range(ncols)]
    has_txn_type = 'transaction type' in header_vals
    if has_txn_type:
        col_map = {'cust_raw': 0, 'date': 1, 'txn_type': 2, 'inv_num': 3, 'product': 4, 'amount': 8}
    else:
        col_map = {'cust_raw': 0, 'date': 1, 'inv_num': 2, 'product': 3, 'amount': 6, 'ae': 9 if ncols > 9 else None}
    data = raw.iloc[header_idx + 1:].copy()
    current_customer = ""
    line_items = []
    for _, row in data.iterrows():
        c0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
        date_val = row.iloc[col_map['date']] if col_map['date'] < ncols else None
        has_date = pd.notna(date_val)
        inv_col = col_map['inv_num']
        inv = row.iloc[inv_col] if inv_col < ncols else None
        has_inv = pd.notna(inv) and str(inv).strip() not in ('', '#')
        if has_txn_type:
            txn = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            if has_date and has_inv and txn not in ('Invoice', 'Credit Memo'): continue
        if c0 and not has_date and not has_inv:
            if not c0.startswith('Total for') and c0 not in ('TOTAL', 'total'): current_customer = c0.strip()
            continue
        if has_date and has_inv:
            amt = clean_money(row.iloc[col_map['amount']]) if col_map['amount'] < ncols else 0.0
            ae_col = col_map.get('ae')
            ae_raw = str(row.iloc[ae_col]).strip() if ae_col and ae_col < ncols and pd.notna(row.iloc[ae_col]) else ''
            ae = QB_NAME_MAP.get(ae_raw, ae_raw) if ae_raw and ae_raw != 'Account Executive' else ''
            inv_str = str(int(inv)) if isinstance(inv, (int, float)) else str(inv).replace('.0', '')
            line_items.append({'qb_customer': current_customer, 'qb_date_raw': date_val, 'qb_invoice': inv_str,
                               'qb_key': clean_inv_key(inv_str), 'qb_line_amount': amt, 'qb_ae': ae})
    if not line_items: return pd.DataFrame()
    df = pd.DataFrame(line_items)
    df['qb_date'] = pd.to_datetime(df['qb_date_raw'], dayfirst=True, errors='coerce')
    invoices = df.groupby('qb_key').agg(qb_customer=('qb_customer', 'first'), qb_date=('qb_date', 'first'),
        qb_invoice=('qb_invoice', 'first'), qb_amount=('qb_line_amount', 'sum'), qb_ae=('qb_ae', 'first'),
        qb_line_count=('qb_line_amount', 'count')).reset_index()
    invoices['qb_amount'] = invoices['qb_amount'].round(2)
    return invoices

def build_excel_rep(rep_name, deals, rc, payout_comm, payout_accel, payout_total, n_missing, period_str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Payout"
    hf = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2E75B6')
    df_ = Font(name='Arial', size=10)
    bf = Font(name='Arial', size=10, bold=True)
    pf = Font(name='Arial', size=18, bold=True)
    plf = Font(name='Arial', size=12, bold=True)
    pfill = PatternFill('solid', fgColor='E2EFDA')
    tfill = PatternFill('solid', fgColor='D6E4F0')
    wfill = PatternFill('solid', fgColor='FFC7CE')
    brd = Border(bottom=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'),
                 left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'))
    cf = '#,##0.00'
    pf2 = '0.0%'
    lv = REP_CONFIG.get(rep_name, {}).get('level', '?')
    rg = REP_CONFIG.get(rep_name, {}).get('region', '?')
    ws['A1'] = f"Payout Report - {rep_name}"
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A2'] = f"Level: {lv} | Region: {rg} | Currency: {rc} | Period: {period_str}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True)
    if n_missing > 0:
        ws['A3'] = f"WARNING: {n_missing} of {len(deals)} deals missing booth items"
        ws['A3'].font = Font(name='Arial', size=10, bold=True, color='9C0006')
    row = 5
    for lbl, val in [("PAYOUT", ""), ("Commission", payout_comm), ("Accelerator", payout_accel), (f"TOTAL PAYOUT ({rc})", payout_total)]:
        for ci in range(1, 4): ws.cell(row=row, column=ci).fill = pfill
        ws.cell(row=row, column=1, value=lbl)
        if isinstance(val, (int, float)) and val != "":
            ws.cell(row=row, column=2, value=val).number_format = cf
            ws.cell(row=row, column=3, value=rc)
        ws.cell(row=row, column=1).font = pf if lbl.startswith("TOTAL") else (plf if lbl == "PAYOUT" else df_)
        ws.cell(row=row, column=2).font = pf if lbl.startswith("TOTAL") else df_
        row += 1
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 8
    ws_d = wb.create_sheet("Deals")
    headers = ['Deal', 'Close Q', 'Paid', 'Booth Items', 'Comm %', 'Commission', 'Tier %', 'Accelerator', 'Payout', 'Channel', 'Status']
    for ci, h in enumerate(headers, 1):
        c = ws_d.cell(row=1, column=ci, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center')
    ds = deals.sort_values('paid_raw', ascending=False)
    for ri, (_, d) in enumerate(ds.iterrows(), 2):
        vals = [d['Deal Name'], d['close_quarter'], round(d['paid_raw'],2), round(d['booth_raw'],2),
                d['comm_rate'], round(d['comm_local'],2), d['tier_rate'], round(d['accel_local'],2),
                round(d['payout_local'],2), d['channel'],
                'REFUND' if d.get('is_refund', False) else ('MISSING' if d['booth_missing'] else 'OK')]
        for ci, val in enumerate(vals, 1):
            c = ws_d.cell(row=ri, column=ci, value=val)
            c.font = df_
            c.border = brd
            if ci in (3,4,6,8,9): c.number_format = cf
            elif ci in (5,7): c.number_format = pf2
        if d['booth_missing'] or d.get('is_refund', False):
            for ci in range(1, len(headers)+1): ws_d.cell(row=ri, column=ci).fill = wfill
    tr = len(ds) + 2
    for ci in range(1, len(headers)+1):
        ws_d.cell(row=tr, column=ci).fill = tfill
        ws_d.cell(row=tr, column=ci).font = bf
    ws_d.cell(row=tr, column=1, value="TOTAL")
    ws_d.cell(row=tr, column=3, value=round(ds['paid_raw'].sum(),2)).number_format = cf
    ws_d.cell(row=tr, column=4, value=round(ds['booth_raw'].sum(),2)).number_format = cf
    ws_d.cell(row=tr, column=6, value=round(ds['comm_local'].sum(),2)).number_format = cf
    ws_d.cell(row=tr, column=8, value=round(ds['accel_local'].sum(),2)).number_format = cf
    ws_d.cell(row=tr, column=9, value=round(ds['payout_local'].sum(),2)).number_format = cf
    pr = tr + 2
    ws_d.cell(row=pr, column=1, value=f"PAYOUT ({rc})").font = Font(name='Arial', size=14, bold=True)
    ws_d.cell(row=pr, column=9, value=payout_total).font = Font(name='Arial', size=14, bold=True)
    ws_d.cell(row=pr, column=9).number_format = cf
    for ci in range(1, len(headers)+1): ws_d.cell(row=pr, column=ci).fill = pfill
    ws_d.column_dimensions['A'].width = 55
    ws_d.column_dimensions['B'].width = 12
    for cl in ['C','D','F','H','I']: ws_d.column_dimensions[cl].width = 15
    for cl in ['E','G']: ws_d.column_dimensions[cl].width = 10
    ws_d.column_dimensions['J'].width = 20
    ws_d.column_dimensions['K'].width = 12
    return wb

st.set_page_config(page_title="Commission Calculator", layout="wide")

st.sidebar.title("Commission Calculator")
page = st.sidebar.radio("", ["Dashboard", "Monthly Payout", "Quarterly Review", "Data Quality"])

st.sidebar.markdown("---")
rates = DEFAULT_RATES.copy()
with st.sidebar.expander("Exchange Rates"):
    for c in ("CAD", "AUD", "GBP", "EUR"):
        rates[c] = st.number_input(f"1 {c} = x USD", value=DEFAULT_RATES[c], format="%.4f", key=f"r_{c}")

c1, c2 = st.columns(2)
hs_file = c1.file_uploader("HubSpot Deals CSV", type=['csv'])
qb_file = c2.file_uploader("QuickBooks Sales by Customer Detail", type=['csv', 'xlsx'])

if not hs_file:
    if 'welcome_shown' not in st.session_state:
        st.session_state.welcome_shown = True
        msg = random.choice([
            "hey teya - you keep this whole machine running. don't let anyone forget it.",
            "teya's commission calculator is online. finance has never looked this good.",
            "another day, another flawless payout run. nice work teya.",
            "teya mode: activated. the reps don't know how lucky they are.",
        ])
        ph = st.empty()
        ph.success(msg)
        _time.sleep(3)
        ph.empty()
    st.info("Upload your HubSpot Deals CSV to get started.")
    st.stop()

try:
    df = pd.read_csv(hs_file)
    df = df.replace("(No value)", np.nan)
    if 'Deal ID' in df.columns and 'Record ID' not in df.columns:
        df = df.rename(columns={'Deal ID': 'Record ID'})
    if 'Deal ID.1' in df.columns: df = df.drop(columns=['Deal ID.1'])
    if 'Record ID' in df.columns: df = df.drop_duplicates(subset=['Record ID'], keep='first')

    df['owner'] = df['Deal owner'].apply(clean_owner)
    df['deal_currency'] = df.get('Currency', pd.Series(['USD'] * len(df))).fillna('USD')
    if 'Paid Total' in df.columns:
        df['paid_raw'] = pd.to_numeric(df['Paid Total'], errors='coerce').fillna(0)
    else:
        df['paid_raw'] = (pd.to_numeric(df.get('Invoice Total inc. tax', 0), errors='coerce').fillna(0)
                          - pd.to_numeric(df.get('Overdue Total', 0), errors='coerce').fillna(0))
        st.sidebar.warning("Paid Total missing - derived from Invoice minus Overdue.")
    df['invoice_raw'] = pd.to_numeric(df['Invoice Total inc. tax'], errors='coerce').fillna(0)
    df['overdue_raw'] = pd.to_numeric(df.get('Overdue Total', 0), errors='coerce').fillna(0)
    df['booth_raw'] = pd.to_numeric(df.get('Booths Items Total Revenue', 0), errors='coerce').fillna(0)
    df['booth_missing'] = (df['booth_raw'] == 0) & ((df['paid_raw'] > 0) | (df['invoice_raw'] > 0))
    df['is_refund'] = (df['paid_raw'] < 0) | (df['invoice_raw'] < 0) | (df['booth_raw'] < 0)

    df['close_date'] = pd.to_datetime(df.get('Close Date'), errors='coerce', utc=True)
    df['close_quarter'] = df['close_date'].apply(lambda x: f"{x.year}-Q{(x.month-1)//3+1}" if pd.notna(x) else "unknown")
    df['hs_paid_date'] = pd.to_datetime(df.get('Latest Invoice Paid Date'), errors='coerce')

    channel_col = None
    for col in df.columns:
        if any(k in col.lower() for k in ('channel', 'source')) and col != 'Transaction Type':
            channel_col = col
            break
    df['channel'] = df[channel_col].fillna('Inbound') if channel_col else 'Inbound'

    df['rep_cfg'] = df['owner'].map(REP_CONFIG)
    df['rep_level'] = df['rep_cfg'].apply(lambda x: x['level'] if isinstance(x, dict) else 'Unknown')
    df['rep_region'] = df['rep_cfg'].apply(lambda x: x['region'] if isinstance(x, dict) else 'Other')
    df['rep_currency'] = df['rep_cfg'].apply(lambda x: x['currency'] if isinstance(x, dict) else 'USD')
    df['region_group'] = df['rep_region'].apply(lambda x: x.split(' ')[0] if pd.notna(x) else 'Other')

    def calc_deal(row):
        dc, rc = row['deal_currency'], row['rep_currency']
        return pd.Series({
            'paid_local': convert_currency(row['paid_raw'], dc, rc, rates),
            'booth_local': convert_currency(row['booth_raw'], dc, rc, rates),
            'booth_usd': convert_currency(row['booth_raw'], dc, 'USD', rates),
            'comm_rate': get_commission_rate(row['owner'], row['channel']),
        })
    calcs = df.apply(calc_deal, axis=1)
    df = pd.concat([df, calcs], axis=1)
    df['comm_local'] = (df['booth_local'] * df['comm_rate']).round(2)

    qb_invoices = pd.DataFrame()
    if qb_file:
        qb_file.seek(0)
        qb_invoices = parse_qb_sales_detail(qb_file)

    df['inv_raw'] = df.get('Invoice Numbers', pd.Series([None] * len(df)))
    hs_inv = df.dropna(subset=['inv_raw']).copy()
    hs_inv['inv_list'] = hs_inv['inv_raw'].astype(str).str.split(',')
    hs_ex = hs_inv.explode('inv_list')
    hs_ex['hs_key'] = hs_ex['inv_list'].str.strip().apply(clean_inv_key)
    hs_ex = hs_ex.drop_duplicates(subset=['hs_key'])

    if len(qb_invoices) > 0:
        inv_date_map = hs_ex[['hs_key', 'Record ID']].merge(
            qb_invoices[['qb_key', 'qb_date']], left_on='hs_key', right_on='qb_key', how='inner')
        if len(inv_date_map) > 0:
            last_inv = inv_date_map.groupby('Record ID')['qb_date'].max().reset_index()
            last_inv.columns = ['Record ID', 'last_invoice_date']
            df = df.merge(last_inv, on='Record ID', how='left')
        else: df['last_invoice_date'] = pd.NaT
    else: df['last_invoice_date'] = pd.NaT

    df['payment_date'] = pd.to_datetime(df['last_invoice_date'].fillna(df['hs_paid_date']), errors='coerce')
    df['payment_month'] = df['payment_date'].apply(lambda x: f"{x.year}-{x.month:02d}" if pd.notna(x) else "unknown")
    df['payment_quarter'] = df['payment_date'].apply(lambda x: f"{x.year}-Q{(x.month-1)//3+1}" if pd.notna(x) else "unknown")

    cqa = df.groupby(['owner', 'close_quarter', 'rep_currency', 'rep_region']).agg(
        cq_booth=('booth_local', 'sum'), cq_deals=('booth_local', 'count')).reset_index()
    cqa['cq_booth'] = cqa['cq_booth'].round(2)
    cqa['tier_name'] = cqa.apply(lambda x: get_tier_name(x['cq_booth'], x['owner']), axis=1)
    cqa['tier_rate'] = cqa.apply(lambda x: _get_tier_rate(x['cq_booth'], x['owner']), axis=1)

    cq_lookup = cqa.set_index(['owner', 'close_quarter'])['tier_rate'].to_dict()
    df['tier_rate'] = df.apply(lambda r: cq_lookup.get((r['owner'], r.get('close_quarter', '')), 0.0), axis=1)
    df['accel_local'] = (df['booth_local'] * df['tier_rate']).round(2)
    df['payout_local'] = (df['comm_local'] + df['accel_local']).round(2)

    all_data = df.copy()
    ar_deals = df[df.get('Invoice Status', pd.Series([''] * len(df))).isin(['Invoices Synced', 'First Invoice Paid']) | (df['overdue_raw'] > 0)].copy()
    no_pay_date = df[(df.get('Invoice Status', pd.Series([''] * len(df))) == 'All Invoices Paid') & (df['payment_date'].isna())]
    booth_missing_all = all_data[all_data['booth_missing']].copy()

    st.sidebar.markdown("---")
    st.sidebar.subheader("Payment Period")
    pay_months = sorted([m for m in df['payment_month'].unique() if m != "unknown"], reverse=True)
    pay_quarters = sorted([q for q in df['payment_quarter'].unique() if q != "unknown"], reverse=True)
    period_type = st.sidebar.radio("Filter By", ["All Dates", "Month", "Quarter"], key="pt")
    period_label = "All Dates"
    if period_type == "Month" and pay_months:
        sel_m = st.sidebar.selectbox("Month", pay_months, key="sm")
        df = df[df['payment_month'] == sel_m]
        period_label = sel_m
    elif period_type == "Quarter" and pay_quarters:
        sel_q = st.sidebar.selectbox("Quarter", pay_quarters, key="sq")
        df = df[df['payment_quarter'] == sel_q]
        period_label = sel_q
    if period_type != "All Dates":
        st.sidebar.caption(f"{len(df)} deals in period")

    known_reps = sorted([r for r in df['owner'].unique() if r in REP_CONFIG])

    if page == "Dashboard":
        st.header("Dashboard")

        ad = all_data[all_data['owner'].isin(REP_CONFIG)]
        total_invoiced = round(ad.apply(
            lambda r: convert_currency(r['invoice_raw'], r['deal_currency'], 'USD', rates), axis=1).sum(), 0)
        total_collected = round(ad.apply(
            lambda r: convert_currency(r['paid_raw'], r['deal_currency'], 'USD', rates), axis=1).sum(), 0)
        total_payout = round(ad.apply(
            lambda r: convert_currency(r['payout_local'], r['rep_currency'], 'USD', rates), axis=1).sum(), 0)
        total_outstanding = round(ad.apply(
            lambda r: convert_currency(r['overdue_raw'], r['deal_currency'], 'USD', rates), axis=1).sum(), 0)
        coll_rate = (total_collected / total_invoiced * 100) if total_invoiced > 0 else 0

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Invoiced", f"${total_invoiced:,.0f}")
        m2.metric("Collected", f"${total_collected:,.0f}")
        m3.metric("Total Payout", f"${total_payout:,.0f}")
        m4.metric("Outstanding", f"${total_outstanding:,.0f}")

        st.progress(min(coll_rate / 100, 1.0))
        st.caption(f"Collection Rate: {coll_rate:.1f}% (all figures USD)")

        c1, c2 = st.columns(2)
        with c1:
            reg_rev = ad.groupby('region_group').apply(
                lambda g: round(g.apply(
                    lambda r: convert_currency(r['paid_raw'], r['deal_currency'], 'USD', rates), axis=1).sum(), 0)
            ).reset_index()
            reg_rev.columns = ['Region', 'Revenue']
            reg_rev = reg_rev.sort_values('Revenue', ascending=True)
            fig_r = px.bar(reg_rev, x='Revenue', y='Region', orientation='h', title="Revenue by Region (USD)")
            fig_r.update_layout(height=350, margin=dict(l=10, r=10, t=40, b=10), showlegend=False)
            fig_r.update_traces(marker_color='#2E86C1')
            st.plotly_chart(fig_r, use_container_width=True)
        with c2:
            ch_rev = ad.groupby('channel').apply(
                lambda g: round(g.apply(
                    lambda r: convert_currency(r['paid_raw'], r['deal_currency'], 'USD', rates), axis=1).sum(), 0)
            ).reset_index()
            ch_rev.columns = ['Channel', 'Revenue']
            fig_c = px.pie(ch_rev, values='Revenue', names='Channel', title="Revenue by Channel", hole=0.4)
            fig_c.update_layout(height=350, margin=dict(l=10, r=10, t=40, b=10))
            st.plotly_chart(fig_c, use_container_width=True)

        st.markdown("##### Top Reps by Payout")
        rep_pay = ad.groupby('owner').apply(
            lambda g: pd.Series({
                'Region': g['region_group'].iloc[0],
                'Currency': g['rep_currency'].iloc[0],
                'Payout (Local)': fmt(g['payout_local'].sum(), g['rep_currency'].iloc[0]),
                'Payout (USD)': round(g.apply(
                    lambda r: convert_currency(r['payout_local'], r['rep_currency'], 'USD', rates), axis=1).sum(), 2),
                'Deals': len(g),
                'Missing Booth': int(g['booth_missing'].sum()),
            })
        ).reset_index().rename(columns={'owner': 'Rep'})
        rep_pay = rep_pay.sort_values('Payout (USD)', ascending=False).head(20)
        fig_p = px.bar(rep_pay, x='Rep', y='Payout (USD)', title="Payout by Rep (USD)", color='Region')
        fig_p.update_layout(height=400, margin=dict(l=10, r=10, t=40, b=10), xaxis_tickangle=-45)
        st.plotly_chart(fig_p, use_container_width=True)

        rep_pay['Payout (USD)'] = rep_pay['Payout (USD)'].apply(lambda v: f"${v:,.2f}")
        st.dataframe(rep_pay, use_container_width=True, hide_index=True)

    elif page == "Monthly Payout":
        st.header(f"Payout - {period_label}")

        if len(no_pay_date) > 0:
            st.warning(f"{len(no_pay_date)} paid deals have no payment date and cannot be assigned to a period.")
        if len(booth_missing_all) > 0:
            st.error(f"{len(booth_missing_all)} deals missing booth items across all data. See Data Quality page.")

        regions = sorted(df[df['owner'].isin(REP_CONFIG)]['region_group'].unique())
        for reg in regions:
            rd = df[(df['region_group'] == reg) & (df['owner'].isin(REP_CONFIG))]
            if len(rd) == 0: continue
            cur = rd.iloc[0]['rep_currency']

            st.markdown(f"### {reg} ({cur})")
            rep_rows = []
            for rep in sorted(rd['owner'].unique()):
                s = rd[rd['owner'] == rep]
                cfg = REP_CONFIG[rep]
                comm = round(s['comm_local'].sum(), 2)
                accel = round(s['accel_local'].sum(), 2)
                pay = round(s['payout_local'].sum(), 2)
                miss = int(s['booth_missing'].sum())
                rep_rows.append({'Rep': rep, 'Level': cfg['level'], 'Commission': fmt(comm, cur),
                                 'Accelerator': fmt(accel, cur), 'PAYOUT': fmt(pay, cur),
                                 'Deals': len(s), 'Missing Booth': miss})
            st.dataframe(pd.DataFrame(rep_rows), use_container_width=True, hide_index=True)

            total_pay = round(rd['payout_local'].sum(), 2)
            st.markdown(f"**Region Total: {fmt(total_pay, cur)}**")

        st.markdown("---")
        if st.button("Download All Rep Reports (.zip)"):
            import zipfile
            zbuf = io.BytesIO()
            with zipfile.ZipFile(zbuf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for rep in known_reps:
                    rd = df[df['owner'] == rep]
                    if len(rd) == 0: continue
                    rc = REP_CONFIG[rep]['currency']
                    comm = round(rd['comm_local'].sum(), 2)
                    accel = round(rd['accel_local'].sum(), 2)
                    pay = round(rd['payout_local'].sum(), 2)
                    miss = int(rd['booth_missing'].sum())
                    wb = build_excel_rep(rep, rd, rc, comm, accel, pay, miss, period_label)
                    buf = io.BytesIO()
                    wb.save(buf)
                    zf.writestr(f"{rep.replace(' ', '_').lower()}.xlsx", buf.getvalue())
            zbuf.seek(0)
            st.download_button(f"Save payout_reports_{period_label}.zip", data=zbuf.getvalue(),
                               file_name=f"payout_reports_{period_label}.zip", mime="application/zip")

        sel_rep = st.selectbox("Individual Rep Detail", [""] + known_reps)
        if sel_rep:
            rd = df[df['owner'] == sel_rep]
            rc = REP_CONFIG[sel_rep]['currency']
            comm = round(rd['comm_local'].sum(), 2)
            accel = round(rd['accel_local'].sum(), 2)
            pay = round(rd['payout_local'].sum(), 2)
            miss = int(rd['booth_missing'].sum())

            if miss > 0:
                st.error(f"{miss} of {len(rd)} deals missing booth items. Payout is understated.")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Commission", fmt(comm, rc))
            m2.metric("Accelerator", fmt(accel, rc))
            m3.metric("PAYOUT", fmt(pay, rc))
            m4.metric("Deals", len(rd))

            disp = pd.DataFrame()
            disp['Deal'] = rd['Deal Name'].values
            disp['Close Q'] = rd['close_quarter'].values
            disp['Paid'] = rd.apply(lambda r: fmt(r['paid_raw'], r['deal_currency']), axis=1).values
            disp['Booth'] = rd.apply(lambda r: fmt(r['booth_raw'], r['deal_currency']), axis=1).values
            disp['Comm'] = rd['comm_local'].apply(lambda v: fmt(v, rc)).values
            disp['Accel'] = rd['accel_local'].apply(lambda v: fmt(v, rc)).values
            disp['Payout'] = rd['payout_local'].apply(lambda v: fmt(v, rc)).values
            disp['Channel'] = rd['channel'].values
            disp['Status'] = rd.apply(lambda r: 'REFUND' if r['is_refund'] else ('MISSING' if r['booth_missing'] else 'OK'), axis=1).values
            st.dataframe(disp.sort_values('Payout', ascending=False), use_container_width=True, hide_index=True)

            wb = build_excel_rep(sel_rep, rd, rc, comm, accel, pay, miss, period_label)
            buf = io.BytesIO()
            wb.save(buf)
            st.download_button(f"Download {sel_rep} Report", data=buf.getvalue(),
                               file_name=f"{sel_rep.replace(' ', '_').lower()}_{period_label}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif page == "Quarterly Review":
        st.header("Quarterly Review - Attainment and Accelerator Tiers")

        close_quarters = sorted([q for q in all_data['close_quarter'].unique() if q != "unknown"], reverse=True)
        sel_cq = st.selectbox("Close Quarter", close_quarters)

        if sel_cq:
            cqa_q = cqa[cqa['close_quarter'] == sel_cq]
            if len(cqa_q) > 0:
                for reg in sorted(cqa_q['rep_region'].apply(lambda x: x.split(' ')[0]).unique()):
                    rq = cqa_q[cqa_q['rep_region'].apply(lambda x: x.split(' ')[0]) == reg]
                    cur = rq.iloc[0]['rep_currency']
                    st.markdown(f"### {reg} ({cur})")
                    rows = []
                    for _, r in rq.sort_values('cq_booth', ascending=False).iterrows():
                        rows.append({'Rep': r['owner'], 'Booth Attainment': fmt(r['cq_booth'], cur),
                                     'Tier': r['tier_name'], 'Tier Rate': f"{r['tier_rate']*100:.2f}%",
                                     'Deals': int(r['cq_deals'])})
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            sel_rep_q = st.selectbox("Rep Detail", [""] + sorted(cqa_q['owner'].unique()) if len(cqa_q) > 0 else [""])
            if sel_rep_q:
                rc = REP_CONFIG.get(sel_rep_q, {}).get('currency', 'USD')
                thresholds = get_bonus_thresholds(sel_rep_q)
                rep_cqa = cqa_q[cqa_q['owner'] == sel_rep_q]
                if len(rep_cqa) > 0:
                    booth_val = rep_cqa.iloc[0]['cq_booth']
                    mv = max(thresholds[-1] * 1.3 if thresholds else 100000, booth_val * 1.1)
                    fig = go.Figure(go.Indicator(
                        mode="gauge+number", value=booth_val,
                        title={'text': f"Booth Attainment ({rc})", 'font': {'size': 16}},
                        number={'prefix': CURRENCY_SYMBOLS.get(rc, "$"), 'valueformat': ",.0f"},
                        gauge={'axis': {'range': [0, mv], 'tickvals': thresholds, 'tickformat': ',.0f'},
                               'bar': {'color': "#2E86C1"}}))
                    fig.update_layout(height=350, margin=dict(l=20, r=20, t=60, b=20))
                    st.plotly_chart(fig, use_container_width=True)

    elif page == "Data Quality":
        st.header("Data Quality")
        dq = all_data

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Deals", f"{len(dq):,}")
        m2.metric("Missing Booth Items", f"{len(booth_missing_all):,}")
        m3.metric("Overdue Deals", f"{(dq['overdue_raw'] > 0).sum():,}")
        m4.metric("No Payment Date", f"{len(no_pay_date):,}")

        st.markdown("##### Missing Booth Items by Currency")
        for cur in sorted(booth_missing_all['deal_currency'].unique()):
            s = booth_missing_all[booth_missing_all['deal_currency'] == cur]
            st.caption(f"{cur}: {len(s)} deals, {fmt(s['paid_raw'].sum(), cur)} paid revenue affected")

        st.markdown("##### Missing Booth Items by Rep")
        rep_miss = []
        for owner in sorted(booth_missing_all['owner'].unique()):
            if owner not in REP_CONFIG: continue
            s = booth_missing_all[booth_missing_all['owner'] == owner]
            cur = REP_CONFIG[owner]['currency']
            total = len(dq[dq['owner'] == owner])
            rep_miss.append({'Rep': owner, 'Currency': cur, 'Missing': len(s),
                             'Total Deals': total, 'Revenue at Risk': fmt(s['paid_raw'].sum(), cur)})
        if rep_miss:
            st.dataframe(pd.DataFrame(rep_miss).sort_values('Missing', ascending=False),
                         use_container_width=True, hide_index=True)

        st.markdown("##### Deal List - Missing Booth Items")
        if len(booth_missing_all) > 0:
            bl = booth_missing_all[['Deal Name', 'owner', 'deal_currency', 'paid_raw', 'invoice_raw', 'close_quarter']].copy()
            bl['Paid'] = bl.apply(lambda r: fmt(r['paid_raw'], r['deal_currency']), axis=1)
            bl['Invoice'] = bl.apply(lambda r: fmt(r['invoice_raw'], r['deal_currency']), axis=1)
            bl = bl.rename(columns={'owner': 'Rep', 'deal_currency': 'Currency', 'close_quarter': 'Close Q'})
            bl = bl[['Deal Name', 'Rep', 'Currency', 'Paid', 'Invoice', 'Close Q']]
            st.dataframe(bl, use_container_width=True, hide_index=True)
            st.download_button("Download Missing Booth Items CSV",
                               data=booth_missing_all[['Record ID', 'Deal Name', 'owner', 'deal_currency',
                                                        'paid_raw', 'invoice_raw', 'close_quarter']].to_csv(index=False),
                               file_name="missing_booth_items.csv", mime="text/csv")

        if len(ar_deals) > 0:
            st.markdown("##### Accounts Receivable")
            ar_disp = ar_deals[['Deal Name', 'owner', 'deal_currency', 'invoice_raw', 'overdue_raw', 'close_quarter']].copy()
            ar_disp['Invoice'] = ar_disp.apply(lambda r: fmt(r['invoice_raw'], r['deal_currency']), axis=1)
            ar_disp['Overdue'] = ar_disp.apply(lambda r: fmt(r['overdue_raw'], r['deal_currency']), axis=1)
            ar_disp = ar_disp.rename(columns={'owner': 'Rep', 'deal_currency': 'Currency', 'close_quarter': 'Close Q'})
            ar_disp = ar_disp[['Deal Name', 'Rep', 'Currency', 'Invoice', 'Overdue', 'Close Q']]
            st.dataframe(ar_disp.sort_values('Overdue', ascending=False), use_container_width=True, hide_index=True)

except Exception as e:
    st.error(f"Error: {e}")
    st.exception(e)
